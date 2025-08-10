import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import shutil
from openpyxl.utils import get_column_letter
import re
import tkinter as tk
import sys
import os
from tkinter import filedialog

if getattr(sys, 'frozen', False):
    script_dir = os.path.dirname(sys.executable)
else:
    script_dir = os.path.dirname(os.path.abspath(__file__))

code_file_path = os.path.join(script_dir, "code.xlsx")

# === ファイル選択ポップアップ ===
def select_file():
    root = tk.Tk()
    root.withdraw()  # メインウィンドウを非表示にする
    file_path = filedialog.askopenfilename(
        title="処理するExcelファイルを選択してください",
        filetypes=[("Excel files", "*.xlsm *.xlsx *.xls")]
    )
    return file_path

# === 入出力ファイルパス ===
input_file = select_file()
if not input_file:
    print("ファイルが選択されませんでした。処理を中止します。")
    exit()

output_file = "整理済みファイル.xlsm"  # 拡張子はxlsmで保存（マクロ保持）

# === 元ファイルをそのままコピーしてマクロを維持 ===
shutil.copyfile(input_file, output_file)

# === コピーしたxlsmファイルを開いて編集 ===
wb = load_workbook(output_file, keep_vba=True)  # ← keep_vba がマクロ維持の鍵
ws = wb.worksheets[0]  # シート1を選択
code_wb = load_workbook(code_file_path, data_only=True)
code_ws = code_wb.active

# === pandasでも読み込み（3行目ヘッダー）※ 必要なら ===
df = pd.read_excel(output_file, sheet_name=0, header=2, engine="openpyxl")

# === ウィンドウ枠固定（A4） ===
ws.freeze_panes = "A4"

# === E列の右隣（F列）に列を挿入 ===
insert_col_idx = 6  # F列（=6）
ws.insert_cols(insert_col_idx)

# === ヘッダー（3行目）に「出願年」と記載 ===
ws.cell(row=3, column=insert_col_idx, value="出願年")

# === F4以降に =YEAR(E4) の式を挿入 ===
for row in range(4, ws.max_row + 1):
    formula = f"=YEAR(E{row})"
    ws.cell(row=row, column=insert_col_idx, value=formula)

# === J列の右隣（K列）に列を挿入 ===
insert_col_idx2 = 11  # K列（=11）
ws.insert_cols(insert_col_idx2)

# === ヘッダー（3行目）に「出願人・権利者(名寄せ)」と記載 ===
ws.cell(row=3, column=insert_col_idx2, value="出願人・権利者(名寄せ")

# === J列（10列目）を参照して「株式会社」を除いた文字列をK列に出力 ===
for row in range(4, ws.max_row + 1):
    original_value = ws.cell(row=row, column=10).value  # J列
    if original_value:
        cleaned_value = str(original_value).replace("株式会社", "")
        ws.cell(row=row, column=11, value=cleaned_value)  # K列（11）
    else:
        ws.cell(row=row, column=11, value="")  # 空白処理

# 「請求の範囲(独立請求項)」列を3行目から探す
claims_col = None
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=3, column=col).value
    if header and "請求の範囲(独立請求項)" in str(header):
        claims_col = col
        break

if claims_col is None:
    raise ValueError("列「請求の範囲(独立請求項)」が見つかりません。")

# 処理：4行目以降、【請求項２】～【請求項９９】以降を抽出し、上書き
pattern = re.compile(r"【請求項([２-９]|[１][０-９]|[２-９][０-９])】")

for row in range(4, ws.max_row + 1):
    cell_value = ws.cell(row=row, column=claims_col).value
    if cell_value:
        text = str(cell_value).strip()
        match = pattern.search(text)
        if match:
            new_text = text[match.start():]
            ws.cell(row=row, column=claims_col, value=new_text)
        else:
            ws.cell(row=row, column=claims_col, value="")  # 請求項2以降がなければ空白
    else:
        ws.cell(row=row, column=claims_col, value="")  # 元々空白ならそのまま空白

# 追加するラベル
labels = ["特許", "特開", "JP", "WO", "CN", "US", "EP", "KR", "TW", "DE", "IN", "AU", "FR", "BR"]

insert_start_col = 21  # U列の右隣はV列＝21列目

# まず14列を右にシフト（列挿入）
for _ in range(len(labels)):
    ws.insert_cols(insert_start_col)
row_header = 3
row_data_start = 4
last_row = ws.max_row

for i, label in enumerate(labels):
    col_idx = insert_start_col + i
    col_letter = get_column_letter(col_idx)
    
    # 見出し行（3行目）にラベルを入力
    ws.cell(row=row_header, column=col_idx, value=label)
    
    # 4行目以降にCOUNTIF式を挿入
    for row in range(row_data_start, last_row + 1):
        formula = f'=COUNTIF(U{row}, "*{label}*")'
        ws.cell(row=row, column=col_idx, value=formula)

# AL列の次（=AM列）は列番号38
insert_col_idx = 38
ws.insert_cols(insert_col_idx)

# ヘッダー（3行目）に見出しを追加
ws.cell(row=3, column=insert_col_idx, value="筆頭FIメイングループ")

# データ行（4行目以降）に関数を挿入
for row in range(4, ws.max_row + 1):
    cell_ref = f"AN{row}"
    formula = f'=LEFT({cell_ref}, FIND("/", {cell_ref}) - 1)'
    ws.cell(row=row, column=insert_col_idx, value=formula)

# 「ＩＰＣ(最新)」列を3行目から探す
claims_col2 = None
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=3, column=col).value
    if header and "ＩＰＣ(最新)" in str(header):
        claims_col2 = col
        break

if claims_col2 is None:
    raise ValueError("列「ＩＰＣ(最新)」が見つかりません。")

# --- 「筆頭IPC列」を挿入 ---
desc_col2 = claims_col2 + 1
ws.insert_cols(desc_col2)
ws.cell(row=3, column=desc_col2).value = "筆頭ＩＰＣ"

# データ行（4行目以降）に関数を挿入
for row in range(4, ws.max_row + 1):
    cell_ref = f"AK{row}"
    formula = f'=IF(COUNTIF({cell_ref},"*;*"), LEFT({cell_ref}, FIND(";", {cell_ref}) - 1), {cell_ref})'
    ws.cell(row=row, column=desc_col2, value=formula)

# 「筆頭ＩＰＣ」列を3行目から探す
claims_col3 = None
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=3, column=col).value
    if header and "筆頭ＩＰＣ" in str(header):
        claims_col3 = col
        break

if claims_col3 is None:
    raise ValueError("列「筆頭ＩＰＣ」が見つかりません。")

# --- 「筆頭IPC説明」を挿入 ---
desc_col3 = claims_col3 + 1
ws.insert_cols(desc_col3)
ws.cell(row=3, column=desc_col3).value = "筆頭ＩＰＣ説明"

# === 3行目の背景を黄色に ===
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
for cell in ws[3]:
    if cell.value not in (None, ""):
        cell.fill = yellow_fill
        
# === フィルターを3行目に設定 ===
last_col = ws.max_column
ws.auto_filter.ref = f"A3:{ws.cell(row=3, column=last_col).coordinate}"

# === 「請求の範囲」列の右に評価列を追加 ===
target_col = None
for col in range(1, last_col + 1):
    if ws.cell(row=3, column=col).value == "請求の範囲":
        target_col = col
        break

if not target_col:
    raise ValueError("「請求の範囲」列が見つかりません。")

# 挿入処理
評価列 = datetime.now().strftime("%y%m%d") + "評価"
ws.insert_cols(target_col + 1)
ws.cell(row=3, column=target_col + 1, value=評価列)
ws.cell(row=3, column=target_col + 1).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# 見出し行から「抄録リンク」が含まれる列を探す
url_col_idx = None
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=3, column=col).value
    if val and "抄録リンク" in str(val):
        print(f"見出し候補: '{val}' → 列番号: {col}")
        url_col_idx = col
        break

if url_col_idx:
    col_letter = get_column_letter(url_col_idx)
    for row in range(4, ws.max_row + 1):
        cell = ws.cell(row=row, column=url_col_idx)
        url = cell.value
        if url:
            formula = f'=HYPERLINK("{url}")'
            cell.value = formula

# --- Fタームコード → 説明の辞書を作成（code.xlsxから）---
code_dict = {}
for row in range(2, code_ws.max_row + 1):  # ヘッダーを除く
    code = code_ws.cell(row=row, column=1).value  # コード列
    desc = code_ws.cell(row=row, column=4).value  # 説明列
    if code and desc:
        code_dict[str(code).strip()] = str(desc).strip()

# --- 「Ｆターム(最新)」列を探す ---
header_row = 3
target_col = None
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=header_row, column=col).value
    if header and "Ｆターム(最新)" in str(header):
        target_col = col
        break

if target_col is None:
    raise Exception("「Ｆターム(最新)」の列が見つかりませんでした。")

# --- 説明列を挿入 ---
desc_col = target_col + 1
ws.insert_cols(desc_col)
ws.cell(row=header_row, column=desc_col).value = "Ｆターム説明"

# --- 各行のFタームコードから説明を取得・結合 ---
for row in range(header_row + 1, ws.max_row + 1):
    fterm_str = ws.cell(row=row, column=target_col).value
    if not fterm_str:
        continue

    fterm_str = str(fterm_str)
    codes = set()  # 重複除去
    for part in fterm_str.split(";"):
        part = part.strip()
        if len(part) >= 5:
            codes.add(part[:5])

    desc_list = [code_dict.get(code, "") for code in codes if code in code_dict]
    ws.cell(row=row, column=desc_col).value = "　".join(desc_list)

# === 列幅15、折り返しなしに設定 ===
for col in ws.columns:
    col_letter = col[0].column_letter
    ws.column_dimensions[col_letter].width = 15
    for cell in col:
        cell.alignment = cell.alignment.copy(wrap_text=False)

# === 保存（マクロ付きのまま） ===
wb.save(output_file)

print("整理処理が完了しました（マクロ保持済み）。保存先:", output_file)
